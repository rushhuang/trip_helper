#!/usr/bin/env python3
"""
xlsx_to_json.py
將 2026沖繩家族旅遊.xlsx 轉換為 pwa/data/itinerary.json

欄位對應：
  A = 日期（3/19(四) 格式，往下延伸至 None）
  B = 時間
  C = 地點名稱
  D = Mapcode / 地址 / 電話 / 營業時間 / 🅿️ 五段式資訊
  E = 備注 / 說明
"""

import re
import json
import datetime
import os
import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', '2026沖繩家族旅遊.xlsx')
JSON_PATH = os.path.join(os.path.dirname(__file__), '..', 'pwa', 'data', 'itinerary.json')

# ── 每日顏色 ─────────────────────────────────────────────────────────
DAY_COLORS = {
    '3/19': '#FF6B6B',
    '3/20': '#FFA94D',
    '3/21': '#69DB7C',
    '3/22': '#4DABF7',
    '3/23': '#DA77F2',
    '3/24': '#F783AC',
    '3/25': '#A9E34B',
    '3/26': '#74C0FC',
}

DAY_LABELS = {
    '3/19': '3/19 (四)',
    '3/20': '3/20 (五)',
    '3/21': '3/21 (六)',
    '3/22': '3/22 (日)',
    '3/23': '3/23 (一)',
    '3/24': '3/24 (二)',
    '3/25': '3/25 (三)',
    '3/26': '3/26 (四)',
}

# ── 地點類型偵測 ──────────────────────────────────────────────────────
STAY_KW    = ['hotel', 'besso', 'lieta', 'strata', 'allivio', 'jr九州', 'jrk',
              '酒店', '飯店', '住宿', '民宿']
TRANSPORT_KW = ['orix', '租車', '還車', '機場', 'airport']
SHOP_KW    = ['コストコ', '好市多', 'big express', 'parco', '塩屋',
              '藥妝', '超市', 'daiso', '大創']
SIGHT_KW   = ['城', '岬', '洞', '海灘', '海濱', '濱', '浜', '水族館', '海洋',
              '燈塔', '遺跡', '島', '萬座毛', '公園', '漁港', '市場', 'anchi',
              '王國村', '玉泉']
FOOD_KW    = ['麵', '拉麵', '壽司', '燒肉', '漢堡', '咖啡', 'coffee', 'bakery',
              'パン', '食堂', '魚', '蝦', '餐廳', '茶', 'cafe', 'a&w', 'a＆w',
              '早餐', '午餐', '晚餐', '宵夜', '飯糰', 'shrimp', '關東煮',
              '鳥貴族', 'pizza', '披薩']

def detect_type(name: str) -> str:
    n = name.lower()
    if any(k in n for k in STAY_KW):      return 'stay'
    if any(k in n for k in TRANSPORT_KW): return 'transport'
    if any(k in n for k in SHOP_KW):      return 'shop'
    if any(k in n for k in SIGHT_KW):     return 'sight'
    if any(k in n for k in FOOD_KW):      return 'food'
    return 'food'  # default

# ── D 欄解析 ──────────────────────────────────────────────────────────
# 無效佔位值（保留欄位但不輸出）
# 額外排除：空行被 \s* 跨行吃到的「地址：」「電話：」等 label 殘留
PLACEHOLDER = re.compile(
    r'^[（(]?(請輸入|查詢中|請查詢|TBD|無)|^(地址|電話|營業時間|🅿)|^\s*$',
    re.IGNORECASE
)

def _extract(pattern: str, text: str, flags=0) -> str:
    """從 text 中提取 pattern 後的【同一行】內容，無效佔位回傳空字串。
    所有 pattern 必須用 [ \\t]* 而非 \\s*，以避免跨行匹配。
    """
    m = re.search(pattern, text, flags)
    if not m:
        return ''
    val = m.group(1).strip()
    return '' if PLACEHOLDER.match(val) else val

def parse_d_column(text: str) -> dict:
    if not text:
        return {}
    t = str(text)
    # 關鍵：使用 [ \t]* 而非 \s*，確保不跨行
    return {
        'mapcode': _extract(r'Mapcode[：:][ \t]*([^\n]+)', t, re.IGNORECASE),
        'address': _extract(r'地址[：:][ \t]*([^\n]+)', t),
        'phone':   _extract(r'電話[：:][ \t]*([^\n]+)', t),
        'hours':   _extract(r'營業時間[：:][ \t]*([^\n]+)', t),
        'parking': _extract(r'🅿[️]?[ \t]*([^\n]+)', t),
    }

# ── 停留時間 / 交通時間 解析 ──────────────────────────────────────────
def parse_duration(text: str) -> int | None:
    """從文字中擷取停留時間（分鐘）。"""
    if not text:
        return None
    # 「停留約1.5小時」「停留約1.5hr」「停留1小時」「停留約40分鐘」
    m = re.search(r'停留[約]?\s*([\d.]+)\s*(小時|hr|h|時間)', text, re.IGNORECASE)
    if m:
        return round(float(m.group(1)) * 60)
    m = re.search(r'停留[約]?\s*([\d.]+)\s*(分鐘|分|min)', text, re.IGNORECASE)
    if m:
        return round(float(m.group(1)))
    # 「停留+吃約1.5hr」
    m = re.search(r'停留.*?([\d.]+)\s*(小時|hr|h|時間)', text, re.IGNORECASE)
    if m:
        return round(float(m.group(1)) * 60)
    m = re.search(r'停留.*?([\d.]+)\s*(分鐘|分|min)', text, re.IGNORECASE)
    if m:
        return round(float(m.group(1)))
    return None

def parse_travel_time(text: str) -> int | None:
    """從文字中擷取交通時間（分鐘）。"""
    if not text:
        return None
    # 「開車過來約20分鐘」「開車過來10分鐘」「距離XX約20分鐘」「過來約1小時」
    m = re.search(r'(?:開車|車程|過來|距離\S*)[約]?\s*([\d.]+)\s*(小時|hr|h|時間)', text, re.IGNORECASE)
    if m:
        return round(float(m.group(1)) * 60)
    m = re.search(r'(?:開車|車程|過來|距離\S*)[約]?\s*([\d.]+)\s*(分鐘|分|min)', text, re.IGNORECASE)
    if m:
        return round(float(m.group(1)))
    return None

# ── 時間格式化 ────────────────────────────────────────────────────────
def fmt_time(val) -> str:
    if val is None:
        return ''
    if isinstance(val, datetime.time):
        return f'{val.hour:02d}:{val.minute:02d}'
    s = str(val).strip()
    # "9:30-13:30" → "9:30"
    if re.match(r'^\d{1,2}:\d{2}', s):
        return s.split('-')[0].strip().split('\n')[0].strip()
    # "8:30\n早餐" → "8:30"
    first_line = s.split('\n')[0].strip()
    if re.match(r'^\d{1,2}:\d{2}', first_line):
        return first_line.split('-')[0].strip()
    # 非時間字串（住宿、早餐…）回傳原值
    return s

# ── 地點名稱清理 ──────────────────────────────────────────────────────
SKIP_NAMES = {'', 'tbd', 'none', '早餐', 'tbд'}

def clean_name(raw) -> str:
    if raw is None:
        return ''
    # 多行名稱：第一行為主名稱，後面為副名稱（日文原名）
    lines = [l.strip() for l in str(raw).strip().split('\n') if l.strip()]
    return ' '.join(lines)  # 用空格連接，方便顯示

def is_skip(name: str) -> bool:
    return name.lower().strip() in SKIP_NAMES

# ── 主程式 ────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['沖繩行程表']

    days_dict: dict[str, list] = {}
    current_date = '3/19'
    stop_counter = 0

    for row_idx in range(3, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        col_b = ws.cell(row=row_idx, column=2).value
        col_c = ws.cell(row=row_idx, column=3).value
        col_d = ws.cell(row=row_idx, column=4).value
        col_e = ws.cell(row=row_idx, column=5).value

        # 更新日期（A 欄出現 "3/XX" 時）
        if col_a and isinstance(col_a, str):
            m = re.match(r'(3/\d+)', col_a)
            if m:
                current_date = m.group(1)

        # 跳過無地點名稱的 row
        name = clean_name(col_c)
        if not name or is_skip(name):
            continue

        info   = parse_d_column(col_d)
        note   = str(col_e).strip() if col_e else ''
        d_text = str(col_d).strip() if col_d else ''
        time_s = fmt_time(col_b)

        # 從 D 欄 + E 欄合併文字中解析停留 / 交通時間
        combined = d_text + '\n' + note
        duration    = parse_duration(combined)
        travel_time = parse_travel_time(combined)

        stop_counter += 1
        stop = {
            'id':         f'stop_{stop_counter:03d}',
            'time':       time_s,
            'type':       detect_type(name),
            'name':       name,
            'mapcode':    info.get('mapcode', ''),
            'address':    info.get('address', ''),
            'phone':      info.get('phone',   ''),
            'hours':      info.get('hours',   ''),
            'parking':    info.get('parking', ''),
            'note':       note,
            'lat':        None,
            'lng':        None,
            'duration':   duration,
            'travelTime': travel_time,
        }

        days_dict.setdefault(current_date, []).append(stop)

    # 組合最終結構
    days = []
    for date in DAY_LABELS:
        if date not in days_dict:
            continue
        days.append({
            'date':  date,
            'label': DAY_LABELS[date],
            'color': DAY_COLORS[date],
            'stops': days_dict[date],
        })

    output = {
        'trip': {
            'title': '2026 沖繩家族旅遊',
            'dates': '3/19(四) – 3/26(四)',
        },
        'days': days,
    }

    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    # ── 統計 ──
    total = sum(len(d['stops']) for d in days)
    print(f'✅ 輸出至 {JSON_PATH}')
    print(f'   {len(days)} 天，共 {total} 個地點\n')
    for d in days:
        print(f'   {d["label"]} ({d["color"]})：{len(d["stops"])} stops')

    # 列出缺少地址的地點（geocode.py 無法處理）
    no_addr = [(d["label"], s["name"])
               for d in days for s in d["stops"] if not s["address"]]
    if no_addr:
        print(f'\n⚠️  以下 {len(no_addr)} 個地點缺少地址（geocode 將跳過）：')
        for label, name in no_addr:
            print(f'   [{label}] {name}')

    # ── 時間統計 ──
    print(f'\n⏱  時間統計：')
    for d in days:
        dur_total = sum(s.get('duration') or 0 for s in d['stops'])
        trv_total = sum(s.get('travelTime') or 0 for s in d['stops'])
        dur_count = sum(1 for s in d['stops'] if s.get('duration'))
        trv_count = sum(1 for s in d['stops'] if s.get('travelTime'))
        total_min = dur_total + trv_total
        print(f'   {d["label"]}：'
              f'停留 {dur_total}min（{dur_count}站）+ '
              f'交通 {trv_total}min（{trv_count}站）= '
              f'{total_min // 60}h{total_min % 60:02d}m')

if __name__ == '__main__':
    main()
