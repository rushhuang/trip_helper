#!/usr/bin/env python3
"""
json_to_xlsx.py
將 itinerary.json 轉換回 Excel 格式

用法：
  python3 scripts/json_to_xlsx.py
  python3 scripts/json_to_xlsx.py --json path/to/data.json --output output.xlsx
"""

import json
import os
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DEFAULT_JSON = os.path.join(os.path.dirname(__file__), '..', 'pwa', 'data', 'itinerary.json')


def hex_to_argb(hex_color: str) -> str:
    """Convert '#FF6B6B' to 'FFFF6B6B' (ARGB format for openpyxl).
    Also handles 3-char hex like '#999' → 'FF999999'.
    """
    h = hex_color.lstrip('#')
    if len(h) == 3:
        h = ''.join(c * 2 for c in h)
    return f'FF{h}'


def main():
    parser = argparse.ArgumentParser(description='JSON → xlsx 轉換')
    parser.add_argument('--json', default=DEFAULT_JSON, help='輸入 JSON 路徑')
    parser.add_argument('--output', default=None, help='輸出 xlsx 路徑')
    args = parser.parse_args()

    with open(args.json, 'r', encoding='utf-8') as f:
        data = json.load(f)

    title = data.get('trip', {}).get('title', '行程')
    output_path = args.output or os.path.join(
        os.path.dirname(args.json), f'{title}.xlsx'
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    # Headers (row 1)
    headers = ['日期', '時間', '地點名稱', '詳細資訊', '備注']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Row 2 is empty (same as original format)
    row_idx = 3

    for day in data.get('days', []):
        color_hex = day.get('color', '#CCCCCC')
        fill = PatternFill(
            start_color=hex_to_argb(color_hex),
            end_color=hex_to_argb(color_hex),
            fill_type='solid',
        )

        for i, stop in enumerate(day.get('stops', [])):
            # A: date (first stop only)
            if i == 0:
                cell_a = ws.cell(row=row_idx, column=1,
                                 value=day.get('label', day.get('date', '')))
                cell_a.fill = fill
                cell_a.font = Font(bold=True)

            # B: time
            ws.cell(row=row_idx, column=2, value=stop.get('time', ''))

            # C: name
            ws.cell(row=row_idx, column=3, value=stop.get('name', ''))

            # D: 5-field info
            d_parts = []
            if stop.get('mapcode'):
                d_parts.append(f"Mapcode：{stop['mapcode']}")
            if stop.get('address'):
                d_parts.append(f"地址：{stop['address']}")
            if stop.get('phone'):
                d_parts.append(f"電話：{stop['phone']}")
            if stop.get('hours'):
                d_parts.append(f"營業時間：{stop['hours']}")
            if stop.get('parking'):
                d_parts.append(f"🅿️ {stop['parking']}")
            d_cell = ws.cell(row=row_idx, column=4, value='\n'.join(d_parts))
            d_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # E: note + duration/travelTime
            e_parts = []
            if stop.get('note'):
                e_parts.append(stop['note'])
            if stop.get('duration') and not (
                stop.get('note') and '停留' in stop['note']
            ):
                dur = stop['duration']
                if dur >= 60 and dur % 60 == 0:
                    e_parts.append(f"停留約{dur // 60}小時")
                elif dur >= 60:
                    e_parts.append(f"停留約{dur / 60:.1f}小時")
                else:
                    e_parts.append(f"停留約{dur}分鐘")
            if stop.get('travelTime') and not (
                stop.get('note')
                and ('開車' in stop['note'] or '車程' in stop['note'])
            ):
                e_parts.append(f"開車過來約{stop['travelTime']}分鐘")
            e_cell = ws.cell(row=row_idx, column=5, value='\n'.join(e_parts))
            e_cell.alignment = Alignment(wrap_text=True, vertical='top')

            row_idx += 1

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 35

    wb.save(output_path)

    # Stats
    total = sum(len(d.get('stops', [])) for d in data.get('days', []))
    print(f'✅ 輸出至 {output_path}')
    print(f'   {len(data.get("days", []))} 天，共 {total} 個地點')


if __name__ == '__main__':
    main()
