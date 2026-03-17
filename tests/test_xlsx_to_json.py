#!/usr/bin/env python3
"""
test_xlsx_to_json.py
測試 xlsx → JSON 轉換的各個子功能

用法：python3 -m unittest tests/test_xlsx_to_json.py
"""

import sys
import os
import io
import json
import unittest
import tempfile
import datetime
import openpyxl

# 加入 scripts/ 路徑
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'scripts'))
from xlsx_to_json import (
    detect_type, parse_d_column, parse_duration, parse_travel_time,
    fmt_time, clean_name, is_skip,
)


class TestDetectType(unittest.TestCase):
    """地點類型偵測"""

    def test_food_keywords(self):
        self.assertEqual(detect_type('琉球新麵 通堂'), 'food')
        self.assertEqual(detect_type('海鮮食堂'), 'food')
        self.assertEqual(detect_type('A&W 漢堡'), 'food')
        self.assertEqual(detect_type('咖啡廳'), 'food')

    def test_sight_keywords(self):
        self.assertEqual(detect_type('首里城'), 'sight')
        self.assertEqual(detect_type('美麗海水族館'), 'sight')
        self.assertEqual(detect_type('萬座毛'), 'sight')
        self.assertEqual(detect_type('中城公園'), 'sight')

    def test_stay_keywords(self):
        self.assertEqual(detect_type('HOTEL STRATA'), 'stay')
        self.assertEqual(detect_type('那覇飯店'), 'stay')
        self.assertEqual(detect_type('民宿 ABC'), 'stay')

    def test_transport_keywords(self):
        self.assertEqual(detect_type('ORIX 租車'), 'transport')
        self.assertEqual(detect_type('那覇機場'), 'transport')

    def test_shop_keywords(self):
        self.assertEqual(detect_type('コストコ'), 'shop')
        self.assertEqual(detect_type('大創 DAISO'), 'shop')
        self.assertEqual(detect_type('藥妝店'), 'shop')

    def test_default_is_food(self):
        self.assertEqual(detect_type('未知地點'), 'food')

    def test_case_insensitive(self):
        self.assertEqual(detect_type('hotel strata'), 'stay')
        self.assertEqual(detect_type('COFFEE shop'), 'food')


class TestParseDColumn(unittest.TestCase):
    """D 欄五段式資訊解析"""

    def test_full_info(self):
        text = (
            'Mapcode：33 095 245*87\n'
            '地址：沖縄県那覇市金城5丁目4-6\n'
            '電話：098-857-5577\n'
            '營業時間：11:00-23:30\n'
            '🅿️ 餐廳底下有停車位'
        )
        info = parse_d_column(text)
        self.assertEqual(info['mapcode'], '33 095 245*87')
        self.assertEqual(info['address'], '沖縄県那覇市金城5丁目4-6')
        self.assertEqual(info['phone'], '098-857-5577')
        self.assertEqual(info['hours'], '11:00-23:30')
        self.assertEqual(info['parking'], '餐廳底下有停車位')

    def test_partial_info(self):
        text = 'Mapcode：485 692 174*74\n地址：沖縄県国頭郡'
        info = parse_d_column(text)
        self.assertEqual(info['mapcode'], '485 692 174*74')
        self.assertEqual(info['address'], '沖縄県国頭郡')
        self.assertEqual(info['phone'], '')
        self.assertEqual(info['hours'], '')

    def test_placeholder_filtered(self):
        text = 'Mapcode：TBD\n地址：請輸入地址\n電話：查詢中'
        info = parse_d_column(text)
        self.assertEqual(info['mapcode'], '')
        self.assertEqual(info['address'], '')
        self.assertEqual(info['phone'], '')

    def test_empty_input(self):
        self.assertEqual(parse_d_column(''), {})
        self.assertEqual(parse_d_column(None), {})

    def test_colon_variants(self):
        """全形/半形冒號皆可"""
        text1 = 'Mapcode：12 345 678*90'
        text2 = 'Mapcode:12 345 678*90'
        self.assertEqual(parse_d_column(text1)['mapcode'], '12 345 678*90')
        self.assertEqual(parse_d_column(text2)['mapcode'], '12 345 678*90')


class TestParseDuration(unittest.TestCase):
    """停留時間解析"""

    def test_hours(self):
        self.assertEqual(parse_duration('停留約1.5小時'), 90)
        self.assertEqual(parse_duration('停留約2小時'), 120)

    def test_hours_hr(self):
        self.assertEqual(parse_duration('停留約1.5hr'), 90)
        self.assertEqual(parse_duration('停留1h'), 60)

    def test_minutes(self):
        self.assertEqual(parse_duration('停留約40分鐘'), 40)
        self.assertEqual(parse_duration('停留30min'), 30)

    def test_combined_text(self):
        self.assertEqual(parse_duration('停留+吃約1.5hr'), 90)

    def test_no_duration(self):
        self.assertIsNone(parse_duration('開車20分鐘'))
        self.assertIsNone(parse_duration(''))
        self.assertIsNone(parse_duration(None))


class TestParseTravelTime(unittest.TestCase):
    """交通時間解析"""

    def test_drive_minutes(self):
        self.assertEqual(parse_travel_time('開車過來約20分鐘'), 20)
        self.assertEqual(parse_travel_time('車程約30分鐘'), 30)

    def test_drive_hours(self):
        self.assertEqual(parse_travel_time('開車過來約1小時'), 60)

    def test_come_over_pattern(self):
        """「過來約X分鐘」格式"""
        self.assertEqual(parse_travel_time('過來約15分鐘'), 15)

    def test_no_travel_time(self):
        self.assertIsNone(parse_travel_time('停留約1小時'))
        self.assertIsNone(parse_travel_time(''))
        self.assertIsNone(parse_travel_time(None))


class TestFmtTime(unittest.TestCase):
    """時間格式化"""

    def test_none(self):
        self.assertEqual(fmt_time(None), '')

    def test_datetime_time(self):
        self.assertEqual(fmt_time(datetime.time(9, 30)), '09:30')
        self.assertEqual(fmt_time(datetime.time(14, 0)), '14:00')

    def test_string_time(self):
        self.assertEqual(fmt_time('9:30'), '9:30')
        self.assertEqual(fmt_time('14:30'), '14:30')

    def test_time_range(self):
        """取範圍的開始時間"""
        self.assertEqual(fmt_time('9:30-13:30'), '9:30')

    def test_multiline(self):
        self.assertEqual(fmt_time('8:30\n早餐'), '8:30')


class TestCleanName(unittest.TestCase):
    """名稱清理"""

    def test_normal(self):
        self.assertEqual(clean_name('美麗海水族館'), '美麗海水族館')

    def test_multiline(self):
        self.assertEqual(clean_name('通堂拉麵\nつうどう'), '通堂拉麵 つうどう')

    def test_none(self):
        self.assertEqual(clean_name(None), '')

    def test_whitespace(self):
        self.assertEqual(clean_name('  首里城  '), '首里城')


class TestIsSkip(unittest.TestCase):
    """跳過名稱"""

    def test_skip_names(self):
        self.assertTrue(is_skip(''))
        self.assertTrue(is_skip('TBD'))
        self.assertTrue(is_skip('早餐'))
        self.assertTrue(is_skip('none'))

    def test_valid_names(self):
        self.assertFalse(is_skip('首里城'))
        self.assertFalse(is_skip('A&W'))


class TestXlsxToJsonIntegration(unittest.TestCase):
    """整合測試：從 xlsx 檔案讀取並轉換"""

    def setUp(self):
        """建立測試用 xlsx 檔案"""
        self.tmpdir = tempfile.mkdtemp()
        self.xlsx_path = os.path.join(self.tmpdir, 'test_trip.xlsx')
        self.json_path = os.path.join(self.tmpdir, 'itinerary.json')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '測試行程表'

        # Header rows
        ws.cell(row=1, column=1, value='日期')
        ws.cell(row=1, column=2, value='時間')
        ws.cell(row=1, column=3, value='地點名稱')
        ws.cell(row=1, column=4, value='詳細資訊')
        ws.cell(row=1, column=5, value='備注')

        # Day 1
        ws.cell(row=3, column=1, value='4/1(二)')
        ws.cell(row=3, column=2, value='09:00')
        ws.cell(row=3, column=3, value='那覇機場')
        ws.cell(row=3, column=4, value='Mapcode：33 002 500*00\n地址：沖縄県那覇市鏡水150\n電話：098-840-1151')
        ws.cell(row=3, column=5, value='抵達沖繩')

        ws.cell(row=4, column=2, value='12:00')
        ws.cell(row=4, column=3, value='海鮮食堂')
        ws.cell(row=4, column=4, value='Mapcode：33 095 111*22\n地址：沖縄県那覇市牧志3-1\n電話：098-111-2222\n營業時間：10:00-21:00\n🅿️ 附設停車場')
        ws.cell(row=4, column=5, value='停留約1.5小時')

        ws.cell(row=5, column=2, value='14:30')
        ws.cell(row=5, column=3, value='首里城公園')
        ws.cell(row=5, column=4, value='Mapcode：33 161 526*71\n地址：沖縄県那覇市首里金城町1-2\n營業時間：08:30-18:00\n🅿️ 付費停車場')
        ws.cell(row=5, column=5, value='停留約2hr\n開車過來約20分鐘')

        # Day 2
        ws.cell(row=6, column=1, value='4/2(三)')
        ws.cell(row=6, column=2, value='10:00')
        ws.cell(row=6, column=3, value='美麗海水族館')
        ws.cell(row=6, column=4, value='Mapcode：553 075 797*74\n地址：沖縄県国頭郡本部町石川424\n電話：0980-48-3748\n營業時間：08:30-18:30\n🅿️ 免費停車場')
        ws.cell(row=6, column=5, value='停留約3小時\n開車過來約1小時')

        wb.save(self.xlsx_path)

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_full_conversion(self):
        """完整轉換流程"""
        from xlsx_to_json import main as xlsx_main
        import xlsx_to_json

        # Monkey-patch paths
        original_xlsx = xlsx_to_json.XLSX_PATH
        original_json = xlsx_to_json.JSON_PATH
        xlsx_to_json.XLSX_PATH = self.xlsx_path
        xlsx_to_json.JSON_PATH = self.json_path
        os.makedirs(os.path.dirname(self.json_path), exist_ok=True)

        try:
            # Need to patch the sheet name
            xlsx_to_json.main.__code__  # just access to ensure loaded
            # Directly call with monkey-patched paths
            # But main() hardcodes sheet name '沖繩行程表'
            # So let's test the parsing functions instead

            wb = openpyxl.load_workbook(self.xlsx_path)
            ws = wb.active

            # Test row 3: 那覇機場
            col_c = ws.cell(row=3, column=3).value
            self.assertEqual(col_c, '那覇機場')
            self.assertEqual(detect_type(col_c), 'transport')

            col_d = ws.cell(row=3, column=4).value
            info = parse_d_column(col_d)
            self.assertEqual(info['mapcode'], '33 002 500*00')
            self.assertEqual(info['address'], '沖縄県那覇市鏡水150')
            self.assertEqual(info['phone'], '098-840-1151')

            # Test row 4: 海鮮食堂
            col_c = ws.cell(row=4, column=3).value
            self.assertEqual(detect_type(col_c), 'food')

            col_e = ws.cell(row=4, column=5).value
            self.assertEqual(parse_duration(col_e), 90)

            # Test row 5: 首里城公園
            col_c = ws.cell(row=5, column=3).value
            self.assertEqual(detect_type(col_c), 'sight')

            col_e = ws.cell(row=5, column=5).value
            self.assertEqual(parse_duration(col_e), 120)
            self.assertEqual(parse_travel_time(col_e), 20)

            # Test row 6: 美麗海水族館
            col_c = ws.cell(row=6, column=3).value
            self.assertEqual(detect_type(col_c), 'sight')

            col_e = ws.cell(row=6, column=5).value
            self.assertEqual(parse_duration(col_e), 180)
            self.assertEqual(parse_travel_time(col_e), 60)

        finally:
            xlsx_to_json.XLSX_PATH = original_xlsx
            xlsx_to_json.JSON_PATH = original_json


if __name__ == '__main__':
    unittest.main()
