#!/usr/bin/env python3
"""
test_roundtrip.py
測試 JSON → xlsx → JSON 往返轉換，驗證資料保真度

用法：python3 -m unittest tests/test_roundtrip.py
"""

import sys
import os
import json
import unittest
import tempfile
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'scripts'))
from json_to_xlsx import main as json_to_xlsx_main
from xlsx_to_json import (
    parse_d_column, parse_duration, parse_travel_time,
    detect_type, fmt_time, clean_name,
)

# ── 原始測試資料 ──────────────────────────────────────────────────────
ORIGINAL = {
    'trip': {
        'title': '沖繩測試旅遊',
        'dates': '4/1(二) – 4/2(三)',
    },
    'days': [
        {
            'date': '4/1',
            'label': '4/1 (二)',
            'color': '#FF6B6B',
            'stops': [
                {
                    'id': 'stop_001',
                    'time': '09:00',
                    'type': 'transport',
                    'name': '那覇機場',
                    'mapcode': '33 002 500*00',
                    'address': '沖縄県那覇市鏡水150',
                    'phone': '098-840-1151',
                    'hours': '24小時',
                    'parking': '付費停車場',
                    'note': '抵達沖繩',
                    'lat': 26.2078,
                    'lng': 127.6462,
                    'duration': None,
                    'travelTime': None,
                },
                {
                    'id': 'stop_002',
                    'time': '12:00',
                    'type': 'food',
                    'name': '通堂拉麵 小祿本店',
                    'mapcode': '33 095 245*87',
                    'address': '沖縄県那覇市金城5丁目4-6',
                    'phone': '098-857-5577',
                    'hours': '11:00-23:30',
                    'parking': '底下有停車位',
                    'note': '',
                    'lat': 26.2120,
                    'lng': 127.6814,
                    'duration': 90,
                    'travelTime': 25,
                },
            ],
        },
        {
            'date': '4/2',
            'label': '4/2 (三)',
            'color': '#FFA94D',
            'stops': [
                {
                    'id': 'stop_003',
                    'time': '10:00',
                    'type': 'sight',
                    'name': '美麗海水族館',
                    'mapcode': '553 075 797*74',
                    'address': '沖縄県国頭郡本部町石川424',
                    'phone': '0980-48-3748',
                    'hours': '08:30-18:30',
                    'parking': '免費停車場',
                    'note': '必看海豚秀',
                    'lat': 26.6934,
                    'lng': 127.8774,
                    'duration': 180,
                    'travelTime': 60,
                },
            ],
        },
    ],
}


class TestRoundTrip(unittest.TestCase):
    """JSON → xlsx → 逐列解析 → 驗證資料保真"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.json_path = os.path.join(self.tmpdir, 'original.json')
        self.xlsx_path = os.path.join(self.tmpdir, 'output.xlsx')

        with open(self.json_path, 'w', encoding='utf-8') as f:
            json.dump(ORIGINAL, f, ensure_ascii=False)

        # Step 1: JSON → xlsx
        sys.argv = ['json_to_xlsx.py', '--json', self.json_path, '--output', self.xlsx_path]
        json_to_xlsx_main()

        # Step 2: Read back xlsx and manually parse (simulate xlsx_to_json logic)
        self.wb = openpyxl.load_workbook(self.xlsx_path)
        self.ws = self.wb.active

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _parse_row(self, row_idx):
        """用 xlsx_to_json 的函式解析單一列"""
        col_b = self.ws.cell(row=row_idx, column=2).value
        col_c = self.ws.cell(row=row_idx, column=3).value
        col_d = self.ws.cell(row=row_idx, column=4).value
        col_e = self.ws.cell(row=row_idx, column=5).value

        name = clean_name(col_c)
        info = parse_d_column(col_d)
        note = str(col_e).strip() if col_e else ''
        d_text = str(col_d).strip() if col_d else ''
        combined = d_text + '\n' + note

        return {
            'time': fmt_time(col_b),
            'type': detect_type(name),
            'name': name,
            'mapcode': info.get('mapcode', ''),
            'address': info.get('address', ''),
            'phone': info.get('phone', ''),
            'hours': info.get('hours', ''),
            'parking': info.get('parking', ''),
            'note': note,
            'duration': parse_duration(combined),
            'travelTime': parse_travel_time(combined),
        }

    def test_stop1_roundtrip(self):
        """那覇機場"""
        orig = ORIGINAL['days'][0]['stops'][0]
        parsed = self._parse_row(3)

        self.assertEqual(parsed['time'], orig['time'])
        self.assertEqual(parsed['name'], orig['name'])
        self.assertEqual(parsed['type'], orig['type'])
        self.assertEqual(parsed['mapcode'], orig['mapcode'])
        self.assertEqual(parsed['address'], orig['address'])
        self.assertEqual(parsed['phone'], orig['phone'])
        self.assertEqual(parsed['hours'], orig['hours'])
        self.assertEqual(parsed['parking'], orig['parking'])
        self.assertEqual(parsed['note'], orig['note'])

    def test_stop2_roundtrip(self):
        """通堂拉麵"""
        orig = ORIGINAL['days'][0]['stops'][1]
        parsed = self._parse_row(4)

        self.assertEqual(parsed['time'], orig['time'])
        self.assertEqual(parsed['name'], orig['name'])
        self.assertEqual(parsed['type'], orig['type'])
        self.assertEqual(parsed['mapcode'], orig['mapcode'])
        self.assertEqual(parsed['address'], orig['address'])
        self.assertEqual(parsed['phone'], orig['phone'])
        self.assertEqual(parsed['hours'], orig['hours'])
        self.assertEqual(parsed['parking'], orig['parking'])
        # Duration/travel time should survive roundtrip
        self.assertEqual(parsed['duration'], orig['duration'])
        self.assertEqual(parsed['travelTime'], orig['travelTime'])

    def test_stop3_roundtrip(self):
        """美麗海水族館"""
        orig = ORIGINAL['days'][1]['stops'][0]
        parsed = self._parse_row(5)

        self.assertEqual(parsed['time'], orig['time'])
        self.assertEqual(parsed['name'], orig['name'])
        self.assertEqual(parsed['type'], orig['type'])
        self.assertEqual(parsed['mapcode'], orig['mapcode'])
        self.assertEqual(parsed['address'], orig['address'])
        self.assertEqual(parsed['phone'], orig['phone'])
        self.assertEqual(parsed['hours'], orig['hours'])
        self.assertEqual(parsed['parking'], orig['parking'])
        self.assertIn('必看海豚秀', parsed['note'])
        self.assertEqual(parsed['duration'], orig['duration'])
        self.assertEqual(parsed['travelTime'], orig['travelTime'])

    def test_date_labels_preserved(self):
        """日期標籤"""
        self.assertEqual(self.ws.cell(row=3, column=1).value, '4/1 (二)')
        self.assertEqual(self.ws.cell(row=5, column=1).value, '4/2 (三)')

    def test_all_stops_present(self):
        """所有站點皆存在"""
        names = []
        for r in range(3, 6):
            v = self.ws.cell(row=r, column=3).value
            if v:
                names.append(v)
        self.assertEqual(len(names), 3)


class TestRoundTripDurationEdge(unittest.TestCase):
    """停留/交通時間往返的邊界情況"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _roundtrip_stop(self, stop_data):
        """單一 stop 的往返測試"""
        data = {
            'trip': {'title': '邊界測試', 'dates': '4/1'},
            'days': [{
                'date': '4/1', 'label': '4/1', 'color': '#999',
                'stops': [stop_data],
            }],
        }
        json_path = os.path.join(self.tmpdir, 'test.json')
        xlsx_path = os.path.join(self.tmpdir, 'test.xlsx')

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)

        sys.argv = ['json_to_xlsx.py', '--json', json_path, '--output', xlsx_path]
        json_to_xlsx_main()

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        col_d = ws.cell(row=3, column=4).value
        col_e = ws.cell(row=3, column=5).value
        d_text = str(col_d).strip() if col_d else ''
        e_text = str(col_e).strip() if col_e else ''
        combined = d_text + '\n' + e_text

        return {
            'duration': parse_duration(combined),
            'travelTime': parse_travel_time(combined),
        }

    def _make_stop(self, **kwargs):
        base = {
            'id': 's1', 'time': '10:00', 'type': 'sight',
            'name': '測試', 'mapcode': '', 'address': '',
            'phone': '', 'hours': '', 'parking': '',
            'note': '', 'lat': None, 'lng': None,
            'duration': None, 'travelTime': None,
        }
        base.update(kwargs)
        return base

    def test_duration_only(self):
        result = self._roundtrip_stop(self._make_stop(duration=60))
        self.assertEqual(result['duration'], 60)

    def test_travel_time_only(self):
        result = self._roundtrip_stop(self._make_stop(travelTime=15))
        self.assertEqual(result['travelTime'], 15)

    def test_both_present(self):
        result = self._roundtrip_stop(self._make_stop(duration=120, travelTime=45))
        self.assertEqual(result['duration'], 120)
        self.assertEqual(result['travelTime'], 45)

    def test_none_values(self):
        result = self._roundtrip_stop(self._make_stop())
        self.assertIsNone(result['duration'])
        self.assertIsNone(result['travelTime'])

    def test_note_with_duration_no_duplicate(self):
        """note 已含「停留」文字時不重複"""
        result = self._roundtrip_stop(
            self._make_stop(note='停留約1小時看風景', duration=60)
        )
        self.assertEqual(result['duration'], 60)


if __name__ == '__main__':
    unittest.main()
