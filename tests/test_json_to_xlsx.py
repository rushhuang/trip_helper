#!/usr/bin/env python3
"""
test_json_to_xlsx.py
測試 JSON → xlsx 轉換

用法：python3 -m unittest tests/test_json_to_xlsx.py
"""

import sys
import os
import json
import unittest
import tempfile
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'scripts'))
from json_to_xlsx import main as json_to_xlsx_main, hex_to_argb

# ── 測試資料 ──────────────────────────────────────────────────────────
SAMPLE_DATA = {
    'trip': {
        'title': '測試旅遊行程',
        'dates': '4/1(二) – 4/3(四)',
    },
    'days': [
        {
            'date': '4/1',
            'label': '4/1 (二)',
            'color': '#FF6B6B',
            'stops': [
                {
                    'id': 'stop_001',
                    'time': '09:00',
                    'type': 'transport',
                    'name': '那覇機場',
                    'mapcode': '33 002 500*00',
                    'address': '沖縄県那覇市鏡水150',
                    'phone': '098-840-1151',
                    'hours': '',
                    'parking': '',
                    'note': '抵達沖繩',
                    'lat': 26.2078,
                    'lng': 127.6462,
                    'duration': None,
                    'travelTime': None,
                },
                {
                    'id': 'stop_002',
                    'time': '12:00',
                    'type': 'food',
                    'name': '海鮮食堂',
                    'mapcode': '33 095 111*22',
                    'address': '沖縄県那覇市牧志3-1',
                    'phone': '098-111-2222',
                    'hours': '10:00-21:00',
                    'parking': '附設停車場',
                    'note': '',
                    'lat': 26.2150,
                    'lng': 127.6830,
                    'duration': 90,
                    'travelTime': 20,
                },
            ],
        },
        {
            'date': '4/2',
            'label': '4/2 (三)',
            'color': '#FFA94D',
            'stops': [
                {
                    'id': 'stop_003',
                    'time': '10:00',
                    'type': 'sight',
                    'name': '首里城公園',
                    'mapcode': '33 161 526*71',
                    'address': '沖縄県那覇市首里金城町1-2',
                    'phone': '',
                    'hours': '08:30-18:00',
                    'parking': '付費停車場',
                    'note': '世界遺產',
                    'lat': 26.2170,
                    'lng': 127.7195,
                    'duration': 120,
                    'travelTime': 30,
                },
            ],
        },
    ],
}


class TestHexToArgb(unittest.TestCase):

    def test_normal(self):
        self.assertEqual(hex_to_argb('#FF6B6B'), 'FFFF6B6B')
        self.assertEqual(hex_to_argb('#4DABF7'), 'FF4DABF7')

    def test_no_hash(self):
        self.assertEqual(hex_to_argb('FFA94D'), 'FFFFA94D')

    def test_three_char_hex(self):
        self.assertEqual(hex_to_argb('#999'), 'FF999999')
        self.assertEqual(hex_to_argb('#F00'), 'FFFF0000')


class TestJsonToXlsxOutput(unittest.TestCase):
    """測試 JSON → xlsx 輸出"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.json_path = os.path.join(self.tmpdir, 'test.json')
        self.xlsx_path = os.path.join(self.tmpdir, 'output.xlsx')

        with open(self.json_path, 'w', encoding='utf-8') as f:
            json.dump(SAMPLE_DATA, f, ensure_ascii=False)

        # Run conversion via CLI arguments
        sys.argv = ['json_to_xlsx.py', '--json', self.json_path, '--output', self.xlsx_path]
        json_to_xlsx_main()
        self.wb = openpyxl.load_workbook(self.xlsx_path)
        self.ws = self.wb.active

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_file_created(self):
        self.assertTrue(os.path.exists(self.xlsx_path))

    def test_sheet_name(self):
        self.assertEqual(self.ws.title, '測試旅遊行程')

    def test_headers(self):
        self.assertEqual(self.ws.cell(row=1, column=1).value, '日期')
        self.assertEqual(self.ws.cell(row=1, column=2).value, '時間')
        self.assertEqual(self.ws.cell(row=1, column=3).value, '地點名稱')
        self.assertEqual(self.ws.cell(row=1, column=4).value, '詳細資訊')
        self.assertEqual(self.ws.cell(row=1, column=5).value, '備注')

    def test_header_bold(self):
        for col in range(1, 6):
            self.assertTrue(self.ws.cell(row=1, column=col).font.bold)

    def test_day1_date_only_first_stop(self):
        """日期只在每日第一站寫入"""
        self.assertEqual(self.ws.cell(row=3, column=1).value, '4/1 (二)')
        # Second stop of day 1 should have no date
        self.assertIsNone(self.ws.cell(row=4, column=1).value)

    def test_day2_date(self):
        self.assertEqual(self.ws.cell(row=5, column=1).value, '4/2 (三)')

    def test_stop_time(self):
        self.assertEqual(self.ws.cell(row=3, column=2).value, '09:00')
        self.assertEqual(self.ws.cell(row=4, column=2).value, '12:00')
        self.assertEqual(self.ws.cell(row=5, column=2).value, '10:00')

    def test_stop_name(self):
        self.assertEqual(self.ws.cell(row=3, column=3).value, '那覇機場')
        self.assertEqual(self.ws.cell(row=4, column=3).value, '海鮮食堂')
        self.assertEqual(self.ws.cell(row=5, column=3).value, '首里城公園')

    def test_d_column_reconstruction(self):
        """D 欄五段式資訊還原"""
        d_val = self.ws.cell(row=3, column=4).value
        self.assertIn('Mapcode：33 002 500*00', d_val)
        self.assertIn('地址：沖縄県那覇市鏡水150', d_val)
        self.assertIn('電話：098-840-1151', d_val)

    def test_d_column_all_five_fields(self):
        """五段全有的情況"""
        d_val = self.ws.cell(row=4, column=4).value
        self.assertIn('Mapcode：33 095 111*22', d_val)
        self.assertIn('地址：沖縄県那覇市牧志3-1', d_val)
        self.assertIn('電話：098-111-2222', d_val)
        self.assertIn('營業時間：10:00-21:00', d_val)
        self.assertIn('🅿️ 附設停車場', d_val)

    def test_d_column_skip_empty(self):
        """空欄位不輸出"""
        d_val = self.ws.cell(row=3, column=4).value
        self.assertNotIn('營業時間', d_val)
        self.assertNotIn('🅿', d_val)

    def test_e_column_note(self):
        """E 欄備注"""
        self.assertEqual(self.ws.cell(row=3, column=5).value, '抵達沖繩')

    def test_e_column_auto_duration(self):
        """E 欄自動補上停留/交通時間"""
        e_val = self.ws.cell(row=4, column=5).value
        self.assertIn('停留約1.5小時', e_val)
        self.assertIn('開車過來約20分鐘', e_val)

    def test_e_column_no_duplicate_duration(self):
        """note 中已有「停留」不重複加"""
        # stop_003 has note='世界遺產', duration=120
        e_val = self.ws.cell(row=5, column=5).value
        self.assertIn('世界遺產', e_val)
        self.assertIn('停留約2小時', e_val)

    def test_day_color_fill(self):
        """日期欄底色"""
        fill = self.ws.cell(row=3, column=1).fill
        self.assertEqual(fill.fgColor.rgb, 'FFFF6B6B')

        fill2 = self.ws.cell(row=5, column=1).fill
        self.assertEqual(fill2.fgColor.rgb, 'FFFFA94D')

    def test_column_widths(self):
        self.assertEqual(self.ws.column_dimensions['A'].width, 14)
        self.assertEqual(self.ws.column_dimensions['B'].width, 10)
        self.assertEqual(self.ws.column_dimensions['C'].width, 28)
        self.assertEqual(self.ws.column_dimensions['D'].width, 45)
        self.assertEqual(self.ws.column_dimensions['E'].width, 35)

    def test_total_data_rows(self):
        """3 個 stop = 3 行資料（row 3-5）"""
        self.assertIsNotNone(self.ws.cell(row=5, column=3).value)
        self.assertIsNone(self.ws.cell(row=6, column=3).value)


class TestJsonToXlsxEdgeCases(unittest.TestCase):
    """邊界測試"""

    def _convert(self, data):
        tmpdir = tempfile.mkdtemp()
        json_path = os.path.join(tmpdir, 'test.json')
        xlsx_path = os.path.join(tmpdir, 'output.xlsx')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
        sys.argv = ['json_to_xlsx.py', '--json', json_path, '--output', xlsx_path]
        json_to_xlsx_main()
        wb = openpyxl.load_workbook(xlsx_path)
        import shutil
        self._tmpdir = tmpdir
        return wb.active

    def tearDown(self):
        import shutil
        if hasattr(self, '_tmpdir'):
            shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_empty_trip(self):
        """空行程不報錯"""
        ws = self._convert({
            'trip': {'title': '空行程', 'dates': ''},
            'days': [],
        })
        self.assertEqual(ws.cell(row=1, column=1).value, '日期')
        self.assertIsNone(ws.cell(row=3, column=1).value)

    def test_stop_without_optional_fields(self):
        """所有選填欄位皆空"""
        ws = self._convert({
            'trip': {'title': '極簡', 'dates': '4/1'},
            'days': [{
                'date': '4/1', 'label': '4/1', 'color': '#999999',
                'stops': [{
                    'id': 's1', 'time': '10:00', 'type': 'sight',
                    'name': '測試景點',
                    'mapcode': '', 'address': '', 'phone': '',
                    'hours': '', 'parking': '', 'note': '',
                    'lat': None, 'lng': None,
                    'duration': None, 'travelTime': None,
                }],
            }],
        })
        self.assertEqual(ws.cell(row=3, column=3).value, '測試景點')
        # D column: no fields → empty string or None (openpyxl returns None for empty)
        d_val = ws.cell(row=3, column=4).value
        self.assertFalse(d_val)  # None or '' both falsy

    def test_long_sheet_name_truncated(self):
        """超過 31 字元的標題截斷"""
        ws = self._convert({
            'trip': {'title': 'A' * 50, 'dates': ''},
            'days': [],
        })
        self.assertEqual(len(ws.title), 31)


if __name__ == '__main__':
    unittest.main()
